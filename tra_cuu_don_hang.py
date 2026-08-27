"""
tra_cuu_don_hang.py
-------------------
Tra cứu hàng loạt thông tin đơn hàng GHN qua API nội bộ nhanh-api.ghn.vn.
Phiên bản hỗ trợ ĐA NGƯỜI DÙNG (Multi-User Safe):
- Thread-safe Session Pool
- Khóa Mutex chống xung đột đăng nhập đồng thời
- Bộ nhớ đệm TTL Cache (60s) tránh gọi trùng đơn
- Tự động giãn cách và xử lý Rate Limit 429 của GHN
- Giao diện Web & Telegram Bot
"""

import os
import sys
import json
import time
import asyncio
import threading
import requests
from requests.adapters import HTTPAdapter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
from collections import Counter
from dotenv import load_dotenv

# Đảm bảo hiển thị tốt tiếng Việt trong console Windows
try:
    if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure'): sys.stderr.reconfigure(encoding='utf-8')
except Exception: pass

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
SESSION_FILE = os.path.join(BASE_DIR, "ghn_session.json")
load_dotenv(os.path.join(BASE_DIR, ".env"))
TELEGRAM_TOKEN   = os.getenv("TELEGRAM_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "-5256491139")
API_BASE    = "https://nhanh-api.ghn.vn"
CONCURRENCY = 3  # Mức concurrency tối ưu với GHN Rate Limit để đạt 100% thành công
VN_TZ       = timezone(timedelta(hours=7))

# Thread-local storage cho session an toàn đa luồng
_thread_local = threading.local()

# Khóa đồng bộ đăng nhập đa người dùng
_login_lock = asyncio.Lock()

# Bộ nhớ đệm kết quả đơn hàng (TTL 60 giây)
_ORDER_CACHE = {}
_CACHE_TTL   = 60  # giây

def get_thread_session(headers):
    """Tạo hoặc lấy session riêng cho mỗi thread."""
    if not hasattr(_thread_local, "session"):
        s = requests.Session()
        s.headers.update(headers)
        adapter = HTTPAdapter(pool_connections=5, pool_maxsize=5, max_retries=1)
        s.mount("https://", adapter)
        s.mount("http://", adapter)
        _thread_local.session = s
    else:
        _thread_local.session.headers.update(headers)
    return _thread_local.session

# Bảng dịch trạng thái GHN sang tiếng Việt chuẩn
STATUS_VI = {
    # Giao hàng
    "DELIVERED":         "Giao thành công",
    "DELIVERING":        "Đang giao hàng",
    "ON_TRIP":           "Đang trên xe giao",
    "READY_TO_DELIVER":  "Sẵn sàng giao",
    "WAITING_DELIVER":   "Chờ giao lại",
    "DELIVER_FAILED":    "Giao thất bại",
    "DELIVERY_FAIL":     "Giao thất bại",
    
    # Lấy hàng
    "READY_TO_PICK":     "Chờ lấy hàng",
    "PICKING":           "Đang đi lấy",
    "PICKED":            "Đã lấy hàng",
    "PICK_FAILED":       "Lấy thất bại",
    
    # Chuyển trả & Trả hàng
    "RETURNED":          "Đã trả hàng",
    "RETURNING":         "Đang chuyển trả",
    "RETURN_DELIVER":    "Đang đi trả hàng",
    "WAITING_TO_RETURN": "Chờ chuyển trả",
    "RETURN_FAILED":     "Trả thất bại",
    "RETURN_FAIL":       "Trả thất bại",
    
    # Kho & Luân chuyển
    "STORING":           "Đang lưu kho",
    "STORED":            "Đang trong kho",
    "TRANSITING":        "Đang luân chuyển",
    "IN_TRANSIT":        "Đang luân chuyển",
    "WAITING_TRANSIT":   "Chờ luân chuyển",
    
    # Khác
    "CANCEL":            "Đã hủy đơn",
    "CANCELLED":         "Đã hủy đơn",
    "LOST":              "Thất lạc / Mất hàng",
    "DAMAGE":            "Hàng hư hỏng",
    "EXCEPTION":         "Sự cố bất thường",
    "WAITING_APPROVAL":  "Chờ duyệt",
}

STATUS_ICONS = {
    "DELIVERED":         "✅",
    "DELIVERING":        "🛵",
    "ON_TRIP":           "🚛",
    "READY_TO_DELIVER":  "📦",
    "WAITING_DELIVER":   "⏳",
    "DELIVER_FAILED":    "❌",
    "DELIVERY_FAIL":     "❌",
    "READY_TO_PICK":     "📋",
    "PICKING":           "🛵",
    "PICKED":            "📦",
    "PICK_FAILED":       "⚠️",
    "RETURNED":          "↩️",
    "RETURNING":         "🔄",
    "RETURN_DELIVER":    "🛵",
    "WAITING_TO_RETURN": "⏳",
    "RETURN_FAILED":     "⚠️",
    "RETURN_FAIL":       "⚠️",
    "STORING":           "🏭",
    "STORED":            "🏭",
    "TRANSITING":        "🚚",
    "IN_TRANSIT":        "🚚",
    "WAITING_TRANSIT":   "⏳",
    "CANCEL":            "🚫",
    "CANCELLED":         "🚫",
    "LOST":              "⛔",
    "DAMAGE":            "💥",
    "EXCEPTION":         "⚠️",
}

ACTION_VI = {
    "RETURN":            "Trả hàng",
    "DELIVER":           "Giao hàng",
    "PICK":              "Lấy hàng",
    "RETURN_IN_TRIP":    "Đang trả (trên xe)",
    "TRANSIT":           "Luân chuyển",
    "WAITING_TO_RETURN": "Chờ chuyển trả",
    "SCAN_TO_STORING":   "Nhập kho chờ trả",
    "PACKED_TO_SORTING": "Đóng gói phân loại",
    "FORCE_RETURN":      "Chuyển trả bắt buộc",
}

STATUS_COLORS = {
    "DELIVERED":        "C6EFCE",  # Xanh lá (thành công)
    "RETURNED":         "FCE4D6",  # Cam nhạt (đã trả)
    "RETURNING":        "FFF2CC",  # Vàng nhạt (đang trả)
    "ON_TRIP":          "DDEBF7",  # Xanh dương nhạt
    "DELIVERING":       "DDEBF7",  # Xanh dương nhạt
    "STORING":          "E2EFDA",  # Xanh bơ nhạt (lưu kho)
    "STORED":           "E2EFDA",  # Xanh bơ nhạt
    "TRANSITING":       "EAEAEA",  # Xám sáng
    "DELIVER_FAILED":   "FFC7CE",  # Đỏ nhạt (thất bại)
    "DELIVERY_FAIL":    "FFC7CE",
    "RETURN_FAILED":    "FFC7CE",
    "RETURN_FAIL":      "FFC7CE",
    "CANCEL":           "D9D9D9",  # Xám (đã hủy)
    "LOST":             "FF9999",  # Đỏ đậm
}

SERVICE_TYPES_VI = {
    1: "Hàng nhẹ (TMĐT)",
    2: "Hàng nặng / Cồng kềnh",
    3: "Chuyển trả",
}


def load_session_headers():
    """Đọc bearer token từ file session."""
    if not os.path.exists(SESSION_FILE):
        return None
    try:
        with open(SESSION_FILE, "r", encoding="utf-8") as f:
            sd = json.load(f)
            for o in sd.get("origins", []):
                if o.get("origin") == "https://nhanh.ghn.vn":
                    for item in o.get("localStorage", []):
                        if item.get("name") == "SESSION":
                            return {
                                "Content-Type": "application/json",
                                "Authorization": f"Bearer {item.get('value')}"
                            }
    except Exception as e:
        print(f"[TraCuu] Lỗi đọc session: {e}")
    return None
  ]
}
                        }
    except Exception as e:
        print(f"[TrCu] Lỗi đọc session: {e}")
    return None


async def ensure_headers(chat_id=None):
    """Đảm bảo headers API luôn hợp lệ (có khóa Mutex chống xung đột đồng thời)."""
    h = load_session_headers()
    if h:
        try:
            r = requests.post(f"{API_BASE}/api/lastmile/user/my-profile",
                              headers=h, json={}, timeout=6)
            if r.status_code == 200 and r.json().get("code") == 0:
                return h
        except Exception:
            pass

    async with _login_lock:
        # Double-check sau khi nhận lock
        h = load_session_headers()
        if h:
            try:
                r = requests.post(f"{API_BASE}/api/lastmile/user/my-profile",
                                  headers=h, json={}, timeout=6)
                if r.status_code == 200 and r.json().get("code") == 0:
                    return h
            except Exception:
                pass

        print("[TrCu] Session hết hạn, đang tự động đăng nhập làm mới...")
        try:
            import ghn_browser_helper
            ctx = await ghn_browser_helper.get_shared_context()
            pg = await ctx.new_page()
            ok = await ghn_browser_helper.check_and_login(
                pg, ctx, "https://nhanh.ghn.vn/lastmile", chat_id=chat_id)
            if ok:
                await ghn_browser_helper.save_session(ctx)
            await pg.close()
            return load_session_headers()
        except Exception as e:
            print(f"[TrCu] Lỗi làm mới session: {e}")
    return None


def _utc2vn(s):
    """Chuyển thời gian UTC sang định dạng ngày giờ Việt Nam."""
    if not s:
        return ""
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(VN_TZ)
        return dt.strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return s


def _post_with_retry(session, url, json_payload, max_retries=4, timeout=12):
    """Thực hiện POST request có cơ chế xử lý HTTP 429 và thử lại thông minh."""
    for attempt in range(max_retries):
        try:
            r = session.post(url, json=json_payload, timeout=timeout)
            if r.status_code == 200:
                return r
            elif r.status_code == 429:
                # Bị GHN Rate Limit: tự động giãn cách và thử lại
                time.sleep(1.0 + 0.5 * attempt)
                continue
            elif r.status_code == 400:
                return r
        except Exception:
            if attempt == max_retries - 1:
                return None
            time.sleep(0.3 * (attempt + 1))
    return None


def lookup_single_order_sync(code, headers, session=None):
    """Tra cứu chi tiết 1 đơn hàng qua 3 API GHN an toàn đa luồng và có Cache."""
    now = time.time()
    if code in _ORDER_CACHE:
        cached_time, cached_data = _ORDER_CACHE[code]
        if now - cached_time < _CACHE_TTL:
            return dict(cached_data)

    if session is None:
        session = get_thread_session(headers)
        
    r = {
        "order_code": code,
        "status_raw": "",
        "status_vi": "",
        "action_raw": "",
        "action_vi": "",
        "delivery_count": 0,
        "last_fail_reason": "",
        "hub_pick": "",
        "hub_deliver": "",
        "hub_return": "",
        "hub_current": "",
        "order_value": "",
        "weight": "",
        "service_type": "",
        "driver_name": "",
        "driver_phone": "",
        "trip_code": "",
        "trip_hub": "",
        "started_at": "",
        "finished_at": "",
        "error": "",
    }
    
    # 1. API Thông tin chung đơn hàng
    try:
        r1 = _post_with_retry(
            session, f"{API_BASE}/api/lastmile/v2/order/search/common-info",
            {"order_code": code}, max_retries=4, timeout=12)
            
        if r1 and r1.status_code == 200:
            d = r1.json().get("data") or {}
            raw_st = d.get("status", "") or d.get("partnerStatus", "")
            r["status_raw"] = raw_st
            r["status_vi"] = STATUS_VI.get(raw_st.upper(), raw_st)
            
            raw_act = d.get("action", "") or d.get("partnerAction", "")
            r["action_raw"] = raw_act
            r["action_vi"] = ACTION_VI.get(raw_act.upper(), ACTION_VI.get(raw_act, raw_act))
            
            r["hub_pick"] = str(d.get("pickId", "") or "")
            r["hub_deliver"] = str(d.get("deliverId", "") or "")
            r["hub_return"] = str(d.get("returnId", "") or "")
            r["hub_current"] = str(d.get("currentId", "") or "")
            
            val = d.get("orderValue")
            r["order_value"] = f"{val:,.0f} đ" if val is not None else ""
            
            w = d.get("weight")
            r["weight"] = f"{w:,.0f}" if w is not None else ""
            
            svc = d.get("serviceTypeID")
            r["service_type"] = SERVICE_TYPES_VI.get(svc, f"Loại {svc}" if svc else "")
        else:
            r["error"] = "Không tìm thấy dữ liệu"
    except Exception as e:
        r["error"] = str(e)

    # 2. API Thông tin tài xế & chuyến đi
    try:
        r2 = _post_with_retry(
            session, f"{API_BASE}/api/lastmile/delivery/get-delivery-info",
            {"orderCode": code}, max_retries=3, timeout=12)
            
        if r2 and r2.status_code == 200:
            d2 = r2.json().get("data") or {}
            r["driver_name"]  = d2.get("driverName", "") or ""
            r["driver_phone"] = d2.get("driverPhone", "") or ""
            r["trip_code"]    = d2.get("tripCode", "") or ""
            r["trip_hub"]     = str(d2.get("hubId", "") or "")
            r["started_at"]   = _utc2vn(d2.get("startedAt", ""))
            r["finished_at"]  = _utc2vn(d2.get("finishedAt", ""))
    except Exception as e:
        prev = r["error"]
        r["error"] = (prev + "; " if prev else "") + f"Lỗi chuyến: {e}"

    # 3. API POD: Thống kê số lần giao & lý do giao thất bại gần nhất
    try:
        r3 = _post_with_retry(
            session, f"{API_BASE}/api/lastmile/v2/pod/get-all",
            {"order_code": code}, max_retries=3, timeout=10)
            
        if r3 and r3.status_code == 200:
            pods = r3.json().get("data", {}).get("pods", []) or []
            deliver_pods = [p for p in pods if p.get("type") == "DELIVER"]
            
            count = len(deliver_pods)
            raw_st = (r.get("status_raw") or "").upper()
            if count == 0 and (raw_st == "DELIVERED" or "thành công" in r.get("status_vi", "").lower()):
                count = 1
            r["delivery_count"] = count
            
            last_fail = next((p.get("fail_note", "") for p in reversed(deliver_pods) if not p.get("is_succeeded") and p.get("fail_note")), "")
            r["last_fail_reason"] = last_fail
    except Exception:
        pass

    # Lưu vào Cache
    _ORDER_CACHE[code] = (now, dict(r))
    return r


async def lookup_batch_orders(codes, headers, progress_cb=None):
    """Tra cứu danh sách đơn hàng song song với Thread-Safe Pool và Rate Limiter."""
    sem = asyncio.Semaphore(CONCURRENCY)
    loop = asyncio.get_event_loop()
    done = [0]

    async def fetch(code):
        async with sem:
            res = await loop.run_in_executor(
                None, lookup_single_order_sync, code, headers)
            done[0] += 1
            if progress_cb:
                await progress_cb(done[0], len(codes))
            # Giãn cách 60ms để không vượt Rate Limit của GHN
            await asyncio.sleep(0.06)
            return res

    results = await asyncio.gather(*[fetch(c) for c in codes])
    return list(results)


def export_to_excel(results, filename=None):
    """Xuất file Excel đầy đủ tiếng Việt, tô màu theo trạng thái và có sheet tổng hợp."""
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(BASE_DIR, "scratch", f"tra_cuu_don_hang_{ts}.xlsx")
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi Tiết Đơn Hàng"

    hF    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hFill = PatternFill("solid", fgColor="1F4E79")
    hA    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb    = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    hdrs = [
        "STT", "Mã Đơn Hàng", "Trạng Thái", "Số Lần Giao", "Lý Do Thất Bại Gần Nhất",
        "Hành Động", "Loại Dịch Vụ", "Giá Trị Đơn", "Khối Lượng (g)",
        "Bưu Cục Lấy", "Bưu Cục Giao", "Bưu Cục Trả", "Bưu Cục Hiện Tại",
        "Tên Shipper", "SĐT Shipper", "Mã Chuyến Đi",
        "Thời Gian Bắt Đầu", "Thời Gian Kết Thúc", "Ghi Chú / Lỗi"
    ]
    widths = [6, 16, 22, 13, 28, 16, 22, 16, 14, 14, 14, 14, 16, 24, 15, 24, 20, 20, 30]

    ws.row_dimensions[1].height = 32
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(1, ci, h)
        c.font = hF
        c.fill = hFill
        c.alignment = hA
        c.border = tb
        ws.column_dimensions[get_column_letter(ci)].width = w

    la = Alignment(horizontal="left", vertical="center")
    ca = Alignment(horizontal="center", vertical="center")
    ra = Alignment(horizontal="right", vertical="center")

    for ri, row in enumerate(results, 2):
        ws.row_dimensions[ri].height = 20
        raw_status = (row.get("status_raw") or "").upper()
        sc = STATUS_COLORS.get(raw_status, "")
        sf = PatternFill("solid", fgColor=sc) if sc else None
        
        vals = [
            ri - 1,
            row.get("order_code", ""),
            row.get("status_vi", "") or row.get("status_raw", ""),
            row.get("delivery_count", 0),
            row.get("last_fail_reason", ""),
            row.get("action_vi", "") or row.get("action_raw", ""),
            row.get("service_type", ""),
            row.get("order_value", ""),
            row.get("weight", ""),
            row.get("hub_pick", ""),
            row.get("hub_deliver", ""),
            row.get("hub_return", ""),
            row.get("hub_current", ""),
            row.get("driver_name", ""),
            row.get("driver_phone", ""),
            row.get("trip_code", ""),
            row.get("started_at", ""),
            row.get("finished_at", ""),
            row.get("error", ""),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.border = tb
            c.font = Font(name="Calibri", size=10)
            if ci in [1, 4, 10, 11, 12, 13, 15, 17, 18]:
                c.alignment = ca
            elif ci in [8, 9]:
                c.alignment = ra
            else:
                c.alignment = la
            if sf:
                c.fill = sf

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Tổng Hợp Thống Kê")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 16

    ws2.cell(1, 1, "📊 TỔNG HỢP PHÂN LOẠI TRẠNG THÁI").font = Font(bold=True, size=12, name="Calibri", color="1F4E79")
    ws2.cell(3, 1, "Trạng Thái").font = Font(bold=True, name="Calibri", color="FFFFFF")
    ws2.cell(3, 1).fill = hFill
    ws2.cell(3, 1).alignment = ca
    ws2.cell(3, 2, "Số Lượng").font = Font(bold=True, name="Calibri", color="FFFFFF")
    ws2.cell(3, 2).fill = hFill
    ws2.cell(3, 2).alignment = ca
    ws2.cell(3, 3, "Tỉ Lệ (%)").font = Font(bold=True, name="Calibri", color="FFFFFF")
    ws2.cell(3, 3).fill = hFill
    ws2.cell(3, 3).alignment = ca

    sc2 = Counter(
        (r.get("status_vi") or r.get("status_raw") or "Không xác định")
        for r in results)
    total = len(results)
    ri = 4
    for st, cnt in sc2.most_common():
        ws2.row_dimensions[ri].height = 18
        c1 = ws2.cell(ri, 1, st)
        c2 = ws2.cell(ri, 2, cnt)
        c3 = ws2.cell(ri, 3, f"{cnt / total * 100:.1f}%" if total else "0%")
        c1.border = tb; c2.border = tb; c3.border = tb
        c1.alignment = la; c2.alignment = ca; c3.alignment = ca
        ri += 1

    ws2.row_dimensions[ri].height = 20
    c_tot1 = ws2.cell(ri, 1, "Tổng cộng")
    c_tot2 = ws2.cell(ri, 2, total)
    c_tot3 = ws2.cell(ri, 3, "100%")
    for c in [c_tot1, c_tot2, c_tot3]:
        c.font = Font(bold=True, name="Calibri")
        c.border = tb
        c.fill = PatternFill("solid", fgColor="EDEDED")
    c_tot1.alignment = la; c_tot2.alignment = ca; c_tot3.alignment = ca

    ws2.cell(ri + 2, 1, f"Thời gian tra cứu: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(italic=True, name="Calibri", color="595959")

    wb.save(filename)
    print(f"[TrCu] Đã xuất Excel: {filename}")
    return filename


def _tg_msg(chat_id, text):
    """Gửi tin nhắn HTML tới Telegram."""
    if not TELEGRAM_TOKEN:
        return
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"}, timeout=15)
    except Exception as e:
        print(f"[TrCu] Lỗi gửi tin nhắn Telegram: {e}")


def _tg_doc(chat_id, path, caption=""):
    """Gửi file tài liệu tới Telegram."""
    if not TELEGRAM_TOKEN:
        return
    try:
        with open(path, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
                data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
                files={"document": f}, timeout=60)
    except Exception as e:
        print(f"[TrCu] Lỗi gửi file Telegram: {e}")


def build_summary_text(results):
    """Tạo tin nhắn tổng hợp kết quả tiếng Việt có dấu chuẩn đẹp."""
    total = len(results)
    errors = [r for r in results if r.get("error") and not r.get("status_raw")]
    ok_results = [r for r in results if not (r.get("error") and not r.get("status_raw"))]
    
    sc = Counter(
        (r.get("status_vi") or r.get("status_raw") or "Không xác định")
        for r in ok_results)
    
    status_icon_lookup = {}
    for r in ok_results:
        st_vi = r.get("status_vi") or r.get("status_raw")
        raw_st = (r.get("status_raw") or "").upper()
        if st_vi and raw_st in STATUS_ICONS:
            status_icon_lookup[st_vi] = STATUS_ICONS[raw_st]

    lines = [
        "📦 <b>KẾT QUẢ TRA CỨU ĐƠN HÀNG GHN</b>",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        f"📊 Tổng số mã: <b>{total}</b> đơn",
        f"✅ Thành công: <b>{total - len(errors)}</b> đơn",
    ]
    if errors:
        lines.append(f"❌ Lỗi / Không tìm thấy: <b>{len(errors)}</b> đơn")
        
    if sc:
        lines.append("")
        lines.append("📈 <b>Phân loại theo trạng thái:</b>")
        for st, cnt in sc.most_common():
            pct = f"{cnt / total * 100:.0f}%" if total else "0%"
            icon = status_icon_lookup.get(st, "•")
            lines.append(f"  {icon} {st}: <b>{cnt}</b> ({pct})")
            
    if errors and len(errors) <= 15:
        lines.append("")
        lines.append("⚠️ <b>Danh sách mã lỗi:</b>")
        for r in errors:
            lines.append(f"  • <code>{r['order_code']}</code>: {r.get('error', 'Không tìm thấy dữ liệu')}")
            
    lines.append("")
    lines.append("📁 <i>Chi tiết từng đơn & số lần giao đã được tổng hợp trong file Excel đính kèm.</i>")
    return "\n".join(lines)


async def run_tra_cuu_don_hang(order_codes, chat_id=None):
    """Hàm chính xử lý tra cứu hàng loạt và gửi báo cáo về Telegram."""
    if not order_codes:
        _tg_msg(chat_id or TELEGRAM_CHAT_ID, "⚠️ Không tìm thấy mã đơn hàng hợp lệ để tra cứu.")
        return False
        
    tc = chat_id or TELEGRAM_CHAT_ID
    total = len(order_codes)
    
    _tg_msg(tc,
            f"🔍 <b>Đang tra cứu {total} mã đơn hàng...</b>\n"
            f"<i>Vui lòng đợi trong giây lát, dự kiến {max(3, total // 3 + 2)} giây...</i>")
            
    headers = await ensure_headers(tc)
    if not headers:
        _tg_msg(tc, "❌ Không thể xác thực session với hệ thống GHN. Vui lòng thử lại sau.")
        return False
        
    last = [0]

    async def pcb(done, total_n):
        m = (done // 25) * 25
        if m > last[0] and m < total_n:
            last[0] = m
            _tg_msg(tc, f"⏳ Đã tra cứu được <b>{done}/{total_n}</b> đơn...")

    loop = asyncio.get_event_loop()
    t0 = loop.time()
    results = await lookup_batch_orders(order_codes, headers, pcb)
    elapsed = loop.time() - t0
    print(f"[TrCu] Hoàn thành {total} đơn trong {elapsed:.1f}s")
    
    # 1. Xuất file Excel
    xls = export_to_excel(results)
    
    # 2. Gửi tin nhắn text tổng hợp
    _tg_msg(tc, build_summary_text(results))
    
    # 3. Gửi file Excel đính kèm
    ts = datetime.now().strftime("%H:%M ngày %d/%m/%Y")
    sec = f"{elapsed:.1f}s" if elapsed < 10 else f"{elapsed:.0f}s"
    _tg_doc(tc, xls,
            f"📊 <b>Kết quả tra cứu {total} đơn hàng</b>\n"
            f"⏰ <i>{ts} — Xử lý trong {sec}</i>")
    return True
